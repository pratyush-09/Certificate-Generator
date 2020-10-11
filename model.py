# Importing the Libraries
import cv2 as cv2   #importing opencv
import openpyxl     #importing openpyxl (to read and write excel files)
from PIL import Image  #importing Image package

# Sample Template Path
template_path='./template.png'

# Data Path
data_path='./data.xlsx'

# Text Styling 
name_font_size=8                 # font-size for text (name)
institute_font_size=6            # font-size for text (institute)
font_color=(0,0,0)               # font-color for text
font=cv2.FONT_HERSHEY_PLAIN      # font-style for text

# Dimensions of the Certificate
coordinate_y_adjustment=15       # dimensions of the certificate vertically
coordinate_x_adjustment=7        # dimensions of the certificate horizontally

# Opening the Data file
obj=openpyxl.load_workbook(data_path)    # opening the excel file
sheet=obj.active # loading the active sheet

# Functioning
for i in range(1,11):    # for repeating the steps 10 times (10 entries present in the dataset)
    get_name=sheet.cell(row=i,column=1)    # reading the name cell
    get_institute=sheet.cell(row=i,column=2)    # reading the institute cell
    candidate_name=get_name.value    # extracting name from cell
    institute_name=get_institute.value     # extracting institute from cell
    img=cv2.imread(template_path)    # reading the sample template as IMAGE
    text_size1=cv2.getTextSize(candidate_name,font,name_font_size,10)[0]    # setting dimensions of the text (name)
    text_size2=cv2.getTextSize(institute_name,font,institute_font_size,10)[0]  # setting dimensions of the text (institute)
    text_x1=(img.shape[1]-text_size1[0])/2+coordinate_x_adjustment    # setting position referring to the certificate dimensions
    text_y1=(img.shape[0]+text_size1[1])/2-coordinate_y_adjustment    # setting position referring to the certificate dimensions
    text_x2=(img.shape[1]-text_size2[0])/2+coordinate_x_adjustment    # setting position referring to the certificate dimensions
    text_y2=(img.shape[0]+text_size2[1])/2-coordinate_y_adjustment    # setting position referring to the certificate dimensions
    text_x1=int(text_x1)        # positioning the text (name) horizontally
    text_y1=int(text_y1-230)    # positioning the text (name) vertically
    text_x2=int(text_x2)        # positioning the text (institute) horiontally
    text_y2=int(text_y2+20)     # positioning the text (institute) vertically
    cv2.putText(img,candidate_name,    # putting text (name) in the sample with the assigned styles
           (text_x1,text_y1),
           font,
           name_font_size,
           font_color,10)
    cv2.putText(img,institute_name,    #putting text (institute) in the sample with the assigned styles
           (text_x2,text_y2),
           font,
           institute_font_size,
           font_color,10)
    imgr=cv2.cvtColor(img,cv2.COLOR_RGB2BGR)    # converting color from RGB to BGR (BGR is default but opencv converts it into RGB, so converting back to BGR)
    image=Image.fromarray(imgr)                 # converting the image array into Image
    image.save('./Certificates/'+candidate_name+' certificate'+'.jpeg')     # saving the certificate image in the desired location with desired names